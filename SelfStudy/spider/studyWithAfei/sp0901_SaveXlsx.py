# # # excel 存储
# 要操作excel文档，需要额外的安装包（openpyxl , pandas）
# pip install openpyxl

# 一个excel文件，
# 整个文件叫工作簿
# 有工作表-->底部的sheet1

# # 创建一个excel（xlsx）文件
from openpyxl import Workbook

# 创建工作簿
wb = Workbook()
# 创建工作表(找到对应的工作表，后续内容时写在工作表中的)
ws = wb.active  # 默认工作表

## 在工作表中添加内容
# 方式1：直接指定单元格
ws["A1"] = "name"
ws["B1"] = "age"
ws["A2"] = "zhang"
ws["B2"] = "20"
ws["H14"] = "test"

# 保存文件
wb.save("sp0902_excel.xlsx")

# # 创建新的工作表
# work sheet
ws2 = wb.create_sheet("st 2")
# 写入内容-方式2：通过append添加列表，类似csv的写入
ws2.append(["name", "age"])
ws2.append(["hao", "18"])
ws2.append(["shuhao", "19"])

wb.save("sp0902_excel2.xlsx")
